from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from hr.report_db import query_report_by_date, query_report_by_employee


FONT_NAME = "Times New Roman"
FONT_SIZE = 13

HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
ALT_FILL = "F2F2F2"
BORDER_COLOR = "000000"


def _style_report(ws, title_text: str, last_col: int, last_row: int):
    thin = Side(style="thin", color=BORDER_COLOR)

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_col
    )

    title = ws.cell(1, 1)
    title.value = title_text
    title.font = Font(
        name=FONT_NAME,
        size=16,
        bold=True
    )
    title.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 26

    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(
            name=FONT_NAME,
            size=FONT_SIZE,
            bold=True,
            color=HEADER_FONT
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = Border(
            top=thin,
            bottom=thin,
            left=thin,
            right=thin
        )

    for row_idx in range(3, last_row + 1):
        ws.row_dimensions[row_idx].height = 22

        for col_idx in range(1, last_col + 1):
            cell = ws.cell(row_idx, col_idx)

            cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            cell.border = Border(
                top=thin,
                bottom=thin,
                left=thin,
                right=thin
            )

            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_FILL)

    ws.freeze_panes = "A3"

    if last_row >= 2:
        ws.auto_filter.ref = f"A2:{get_column_letter(last_col)}{last_row}"

    for col_idx in range(1, last_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for row_idx in range(1, last_row + 1):
            value = ws.cell(row_idx, col_idx).value
            max_len = max(max_len, len(str(value or "")))

        ws.column_dimensions[col_letter].width = min(max_len + 4, 28)


def export_total_excel(from_date: str, to_date: str, save_path: str):
    rows = query_report_by_date(from_date, to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao tong hop"

    headers = [
        "Ngày",
        "Mã đơn",
        "Mã NV",
        "Tên nhân viên",
        "Bộ phận",
        "Camera",
        "Thời gian bắt đầu",
        "Thời gian kết thúc",
        "Độ dài (giây)",
        "Dung lượng (MB)",
    ]

    ws.append(["BÁO CÁO TỔNG HỢP"])
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get("date", ""),
            r.get("order_code", ""),
            r.get("employee_id", ""),
            r.get("employee_name", ""),
            r.get("department", ""),
            r.get("camera_name", ""),
            r.get("start_time", ""),
            r.get("end_time", ""),
            r.get("duration_sec", 0),
            r.get("file_size_mb", 0),
        ])

    last_row = ws.max_row
    last_col = len(headers)

    _style_report(ws, "BÁO CÁO TỔNG HỢP", last_col, last_row)

    wb.save(save_path)
    return save_path


def export_employee_excel(from_date: str, to_date: str, save_path: str):
    rows = query_report_by_employee(from_date, to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao nhan vien"

    headers = [
        "Mã NV",
        "Tên nhân viên",
        "Bộ phận",
        "Chức vụ",
        "Tổng đơn",
        "Tổng video",
        "Tổng thời lượng (giây)",
        "Tổng thời lượng (phút)",
        "Tổng dung lượng (MB)",
    ]

    ws.append(["BÁO CÁO THEO NHÂN VIÊN"])
    ws.append(headers)

    for r in rows:
        duration_sec = r.get("total_duration_sec") or 0

        ws.append([
            r.get("employee_id", ""),
            r.get("employee_name", ""),
            r.get("department", ""),
            r.get("position", ""),
            r.get("total_orders", 0),
            r.get("total_videos", 0),
            duration_sec,
            round(duration_sec / 60, 2),
            round(r.get("total_size_mb") or 0, 2),
        ])

    last_row = ws.max_row
    last_col = len(headers)

    _style_report(ws, "BÁO CÁO THEO NHÂN VIÊN", last_col, last_row)

    wb.save(save_path)
    return save_path